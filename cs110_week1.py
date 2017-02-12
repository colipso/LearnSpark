#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Fri Feb 10 19:19:48 2017

@author: hp
"""
from openpyxl import load_workbook
import os
import datetime
import traceback
from pyspark import SparkContext
from pyspark.sql import SQLContext
import seaborn as sbn
from pyspark.ml.feature import VectorAssembler
from pyspark.sql.functions import sum

sc = SparkContext("local[4]","cs110 Lecture App")
sc.setLogLevel("ERROR")
sqc = SQLContext(sc)

isDraw = False


def outPutPrint(*context):
    '''
    for output some infomation
    '''
    outputlogo = "---->" + "[" + str(datetime.datetime.now()) + "]"
    string_print = ""
    for c in context:
        string_print += str(c)+"  "
    content = outputlogo +string_print + '\n'
    f = open("log.txt",'a')
    f.write(content)
    f.close;
    print outputlogo,string_print,'\n'
    return True

def memory_stat():
    '''
    return memory usage. returntype is dict
    '''
    mem = {}  
    f = open("/proc/meminfo")  
    lines = f.readlines()  
    f.close()  
    for line in lines:  
        if len(line) < 2: continue  
        name = line.split(':')[0]  
        var = line.split(':')[1].split()[0]  
        mem[name] = long(var) * 1024.0  
    #mem['MemUsed'] = mem['MemTotal'] - mem['MemFree'] - mem['Buffers'] - mem['Cached']
    memNotUsed_G = "%.2f"% (mem['MemFree'] *1.0/(1024.0*1024.0*1024.0))
    outPutPrint("Free Memory is ",memNotUsed_G , "G")
    return memNotUsed_G 


def clear_env():
    '''
    for prevent memory error,clear python veriable
    '''
    for key in globals().keys():
        if key not in ["memory_stat" , "outPutPrint" , "time" , "datetime"]:
            if not key.startswith("__"):
                globals().pop(key)
    memory_stat()
    outPutPrint("The python env is cleared")
    memory_stat()




def transformXlsx(filename , savefile):
    '''
    transform origian excel file to txt
    '''
    wb = load_workbook(filename = filename)
    sheets = wb.get_sheet_names()
    data = []
    for s in sheets:
        sheetData = []
        sheet = wb.get_sheet_by_name(s)
        maxcolnum = sheet.max_column + 1
        maxrownum = sheet.max_row + 1
        for row in range(2,maxrownum):
            colData = []
            for col in range(1,maxcolnum):
                celldata = sheet.cell(row = row ,column = col).value
                #rint celldata
                colData.append(celldata)
            sheetData.append(colData)
        data.extend(sheetData)
    f = open(savefile,'w')
    for line in data:
        s = ""
        for elem in line:
            s += str(elem) + "\t"
        s = s[:-1] + "\n"
        f.write(s)
    f.close()
    outPutPrint("Excel file transform completed")
    
    
if not os.path.exists('./data/data.txt'):
    transformXlsx('./data/Folds5x2_pp.xlsx' ,'./data/data.txt')
else:
    outPutPrint("transformed data exist")
    

try:
    #read data
    def returnSplitedData(r):
        returnData = []
        for d in r.split('\t'):
            returnData.append(float(d))
        return returnData
    
    powerPlantRDD = (sc
                    .textFile('./data/data.txt')
                    .map(returnSplitedData) )
    powerPlantDF = sqc.createDataFrame(powerPlantRDD,['AT','V','AP','RH','PE'])
    powerPlantDF.show(truncate = False)
    outPutPrint("powerPlantDF's dtype is :" , powerPlantDF.dtypes)
    outPutPrint("count powerPlantDF is ",powerPlantDF.count())
    
    #regist dataframe as table
    tables = sqc.sql('show tables')
    tables.show()
    sqc.sql('drop table if exists power_plant')
    sqc.registerDataFrameAsTable(powerPlantDF , 'power_plant')
    outPutPrint("regist dataframe success")
    
    AT_DF = sqc.sql("select AT from power_plant")
    outPutPrint("sql result type is " , type(AT_DF))
    AT_DF.show()
    describe_DF = sqc.sql("desc power_plant")
    outPutPrint("sql result type is " , type(describe_DF))
    describe_DF.show()
    
    #describe data
    df = sqc.table("power_plant")
    df.describe().show()
    cols = df.columns
    outPutPrint("table cols is " , type(cols))
    #compair AT and PE
    for col in cols:
        print col
        if col != "PE":
            at_pe_DF = sqc.sql("select %s , PE from power_plant" % col)
            at_pe_DF.show()
            if isDraw:
                sbn.jointplot(x = col , y = "PE" ,
                          kind = 'reg',data = at_pe_DF.toPandas())
    
    #Mechina learning
    #create features
    dataSetDF = sqc.table("power_plant")
    vectorizer = VectorAssembler(inputCols = ["AT", "V", "AP", "RH"], 
                                 outputCol = 'features')
    seed = 1800009193L
    (testSetDF , trainSetDF) = dataSetDF.randomSplit([0.2,0.8] ,seed)
    from pyspark.ml.regression import LinearRegression
    from pyspark.ml.regression import LinearRegressionModel
    from pyspark.ml import Pipeline
    lr = LinearRegression(predictionCol = 'Predicted_PE',
                         labelCol = 'PE',
                         maxIter = 100 ,
                         regParam = 0.1)
    outPutPrint("linearRegression func param are:" ,lr.explainParams())
    lrPipeLine = Pipeline()
    lrPipeLine.setStages([vectorizer , lr])
    lrModel = lrPipeLine.fit(trainSetDF)
    #outPutPrint("trained lrModel param are:",lrModel.explainParams())
    intercept = lrModel.stages[1].intercept
    weights = lrModel.stages[1].coefficients
    outPutPrint("lrmodel's intercept is :",intercept)
    outPutPrint("lrmodel's weights is ",weights)
    predictResultDF = lrModel.transform(testSetDF)
    predictResultDF.show()
    if isDraw:
        sbn.jointplot(x = "PE" , y = "Predicted_PE" , 
                  data = predictResultDF.toPandas(),kind = 'reg')
    
    #evaluate model
    from pyspark.ml.evaluation import RegressionEvaluator
    regEval = RegressionEvaluator(predictionCol = "Predicted_PE" ,
                                  labelCol = "PE" , 
                                  metricName = "rmse")
    rmse = regEval.evaluate(predictResultDF)
    outPutPrint("Root mean square error is " , rmse)
    
    r2 = regEval.evaluate(predictResultDF , {regEval.metricName:'r2'})
    outPutPrint("R2 is " , r2)
    
    sqc.sql("drop table if exists Power_Plant_RMSE_Evaluation")
    (predictResultDF
     .selectExpr('PE' ,'Predicted_PE' ,
                               'PE-Predicted_PE Residual_Error',
                               '(PE-Predicted_PE)/%f Within_RSME'%rmse)
    .registerTempTable('Power_Plant_RMSE_Evaluation'))
    
    Power_Plant_RMSE_EvaluationDF = sqc.sql('select * from Power_Plant_RMSE_Evaluation')
    Power_Plant_RMSE_EvaluationDF.show()
    
    rsmeDF = Power_Plant_RMSE_EvaluationDF.select("Within_RSME")
    if isDraw:
        sbn.distplot(a = rsmeDF.toPandas().Within_RSME)
    
    sql_string = "select case when Within_RSME <= 1.0 AND Within_RSME >= -1.0 then 1 \
                              when Within_RSME <= 2.0 AND Within_RSME >= -2.0 then 2 \
                              else 3 \
                         end RSMERange , count(*) as count \
                  from Power_Plant_RMSE_Evaluation \
                  group by case when Within_RSME <= 1.0 AND Within_RSME >= -1.0 then 1 \
                              when Within_RSME <= 2.0 AND Within_RSME >= -2.0 then 2 \
                              else 3 \
                            end"
    RMSERangeDF = sqc.sql(sql_string)
    RMSESUMDF = RMSERangeDF.select(RMSERangeDF['count']).groupBy().sum().collect()
    RMSESUM = RMSESUMDF[0]['sum(count)']
    outPutPrint("RMSE sum is " , RMSESUM)
    RMSERangePercentDF = RMSERangeDF.selectExpr("RSMERange" , "count*1.0/%d percent" % RMSESUM)
    RMSERangePercentDF.show()
    
    #select best model
    from pyspark.ml.tuning import ParamGridBuilder ,CrossValidator
    crossval = CrossValidator(estimator = lrPipeLine ,
                              evaluator = regEval ,
                              numFolds = 3)
    regParam = [x/100.0 for x in range(1,11)]
    paramGrid = (ParamGridBuilder()
    .addGrid(lr.regParam , regParam)
    .build())
    crossval.setEstimatorParamMaps(paramGrid)
    cvModel = crossval.fit(trainSetDF).bestModel
    predictedAndLabelDF = cvModel.transform(testSetDF)
    cvRMSE = regEval.evaluate(predictedAndLabelDF)
    outPutPrint("choosed model RMSE is " , cvRMSE)
                          
    #use decisionTreeRegression
    from pyspark.ml.regression import DecisionTreeRegressor
    dt = DecisionTreeRegressor()
    (dt.setLabelCol('PE')
    .setPredictionCol('Predicted_PE')
    .setFeaturesCol('features')
    .setMaxBins(100))
    dtPipeline = Pipeline()
    dtPipeline.setStages([vectorizer , dt])
    crossval.setEstimator(dtPipeline)
    paramGrid = (ParamGridBuilder().addGrid(dt.maxDepth , [2,3]).build())
    crossval.setEstimatorParamMaps(paramGrid)
    dtModel = crossval.fit(trainSetDF).bestModel
    dtpredictedAndLabelDF = dtModel.transform(testSetDF)
    dtcvRMSE = regEval.evaluate(dtpredictedAndLabelDF)
    outPutPrint("decision tree choosed model RMSE is " , dtcvRMSE)
    
    outPutPrint("describe decision :", dtModel.stages[-1]._java_obj.toDebugString())
    
    #use random forest regressor
    from pyspark.ml.regression import RandomForestRegressor
    rf = RandomForestRegressor()#create rfmodel
    (rf.setLabelCol('PE') #set rf model
    .setPredictionCol('Predicted_PE')
    .setFeaturesCol('features')
    .setSeed(100088121L)
    .setMaxDepth(8)
    .setNumTrees(30))
    rfPipeline = Pipeline() #create pipline
    rfPipeline.setStages([vectorizer , rf]) #create pipeline .ps:vectorizer already created
    crossval.setEstimator(rfPipeline) #set crossval .PS:crossval already created
    rfModel = crossval.fit(trainSetDF).bestModel #train and get model
    rfpredictedAndLabelDF = rfModel.transform(testSetDF) #predict
    rfRMSE = regEval.evaluate(rfpredictedAndLabelDF) #calc rmse
    outPutPrint("random forest regression RMSE is " ,rfRMSE)
    



                     
                          
except Exception as e:
    outPutPrint("[ERROR]",e)
    traceback.print_exc()
finally:
    #time.sleep(30)#for check sparkUI before cluster closed
    sc.stop()
    outPutPrint("[INFO]The spark cluster stopped")
    clear_env()
    
