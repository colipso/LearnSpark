#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Sun Feb 12 12:58:59 2017

@author: hp
"""

#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Fri Feb 10 19:19:48 2017

@author: hp
"""
from openpyxl import load_workbook
import os
import datetime
import traceback
from pyspark import SparkContext
from pyspark.sql import SQLContext
import seaborn as sbn
from pyspark.ml.feature import VectorAssembler
from pyspark.sql.functions import sum , col

sc = SparkContext("local[4]","cs110 Lecture App")
sc.setLogLevel("ERROR")
sqc = SQLContext(sc)

isDraw = False


def outPutPrint(*context):
    '''
    for output some infomation
    '''
    outputlogo = "---->" + "[" + str(datetime.datetime.now()) + "]"
    string_print = ""
    for c in context:
        string_print += str(c)+"  "
    content = outputlogo +string_print + '\n'
    f = open("log.txt",'a')
    f.write(content)
    f.close;
    print outputlogo,string_print,'\n'
    return True

def memory_stat():
    '''
    return memory usage. returntype is dict
    '''
    mem = {}  
    f = open("/proc/meminfo")  
    lines = f.readlines()  
    f.close()  
    for line in lines:  
        if len(line) < 2: continue  
        name = line.split(':')[0]  
        var = line.split(':')[1].split()[0]  
        mem[name] = long(var) * 1024.0  
    #mem['MemUsed'] = mem['MemTotal'] - mem['MemFree'] - mem['Buffers'] - mem['Cached']
    memNotUsed_G = "%.2f"% (mem['MemFree'] *1.0/(1024.0*1024.0*1024.0))
    outPutPrint("Free Memory is ",memNotUsed_G , "G")
    return memNotUsed_G 


def clear_env():
    '''
    for prevent memory error,clear python veriable
    '''
    for key in globals().keys():
        if key not in ["memory_stat" , "outPutPrint" , "time" , "datetime"]:
            if not key.startswith("__"):
                globals().pop(key)
    memory_stat()
    outPutPrint("The python env is cleared")
    memory_stat()




def transformXlsx(filename , savefile):
    '''
    transform origian excel file to txt
    '''
    wb = load_workbook(filename = filename)
    sheets = wb.get_sheet_names()
    data = []
    for s in sheets:
        sheetData = []
        sheet = wb.get_sheet_by_name(s)
        maxcolnum = sheet.max_column + 1
        maxrownum = sheet.max_row + 1
        for row in range(2,maxrownum):
            colData = []
            for col in range(1,maxcolnum):
                celldata = sheet.cell(row = row ,column = col).value
                #rint celldata
                colData.append(celldata)
            sheetData.append(colData)
        data.extend(sheetData)
    f = open(savefile,'w')
    for line in data:
        s = ""
        for elem in line:
            s += str(elem) + "\t"
        s = s[:-1] + "\n"
        f.write(s)
    f.close()
    outPutPrint("Excel file transform completed")
    
    
if not os.path.exists('./data/data.txt'):
    transformXlsx('./data/Folds5x2_pp.xlsx' ,'./data/data.txt')
else:
    outPutPrint("transformed data exist")
    

try:
    #test sum
    data1 = [('Alice', 1), ('Bob', 2), ('Bill', 4)]
    df1 = sqc.createDataFrame(data1,['name','age'])
    a = df1.groupBy().sum().collect()
    
    
    print a[0]['sum(age)']

    
except Exception as e:
    outPutPrint("[ERROR]",e)
    traceback.print_exc()
finally:
    #time.sleep(30)#for check sparkUI before cluster closed
    sc.stop()
    outPutPrint("[INFO]The spark cluster stopped")
    #clear_env()
    
